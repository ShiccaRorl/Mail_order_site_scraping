# -*- coding:utf-8 -*-

# python -m pip install --upgrade pip
# pip install -U sqlalchemy
# pip install -U flake8
# pip install -U autopep8

# pip install -U openpyxl
# pip install -U beautifulsoup4
# pip install -U requests
# pip install -U pyinstaller

# pyinstaller kabu.py --onefile

import os
import glob
import re
import datetime
import time
import subprocess
import argparse
import sys
import logging

from sqlalchemy.ext.automap import automap_base
from sqlalchemy.orm import Session
from sqlalchemy import create_engine

from openpyxl import load_workbook, Workbook

Base = automap_base()

# engine, suppose it has two tables 'user' and 'address' set up
#db_path = "sqlite:///./../../db.sqlite3"
db_path = "sqlite:///./db.sqlite3"
engine = create_engine(db_path)
# 個人情報の塊なのでGITの外に設置

# reflect the tables
Base.prepare(engine, reflect=True)

# mapped classes are now created with names by default
# matching that of the table name.

#seeds = Base.classes.seeds
dir_db = Base.classes.dir_db
Product = Base.classes.Product

#print('//DESKTOP-FK98SN0/Users/Public/Documents/吉本さんPCから移動したファイル/メルカリラクマ出品画像/メルカリ')

#ROOT_PATH = '//DESKTOP-FK98SN0/Users/Public/Documents/吉本さんPCから移動したファイル/メルカリラクマ出品画像/メルカリ'#.encode("cp932").replace("/", "\\")
#ROOT_PATH = '//DESKTOP-FK98SN0/Users/Public/Documents/吉本さんPCから移動したファイル/メルカリラクマ出品画像/メルカリ'
#ROOT_PATH = '//vmware-host/Shared Folders/D/共有/Down'
#print(ROOT_PATH)

# Excel path
excel_path1 = "C:/Users/user/Downloads/バックアップ/プログラム/バックアップ/保存/2020年9月在庫表.xlsx"
excel_path2 = "./2020年9月在庫表.xlsx"

formatter = '%(asctime)s:%(message)s'
logging.basicConfig(filename='test.log', level=logging.DEBUG, format=formatter)




#import openpyxl

excel_path1 = "C:/Users/user/Downloads/バックアップ/プログラム/バックアップ/保存/2020年9月在庫表.xlsx"
excel_path2 = "./2020年9月在庫表.xlsx"
wb = load_workbook(excel_path1)

ws = wb["マスター"]
wb.copy_worksheet(ws)
wb.save("./Excel_db.xlsx")


class Excel_Analysis():
    # エクセル解析
    def excel_slise():
        wb1 = load_workbook(excel_path2)
        ws_copy = wb1.copy_worksheet("マスター")
    
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new = ws_copy
        wb_new.save("./マスター.xlsx")

    def get_excel_db():
        # マスターコピー
        #wb_new = Workbook()
        #ws_new = wb_new.active
        #ws_new.title = "マスター"
    
        wb = load_workbook(excel_path2)
        ws = wb["マスター"]
        print(ws.max_row)
        s = 5
        while s <= ws.max_row:
            print(s)
            print(ws[f"B{s}"].value)
            session = Session(engine)
            if session.query(Product).filter(Product.code == ws[f"B{s}"].value).first() == None:
                print(f"insert {s :} ")
                session.add(Product(code = ws[f"B{s}"].value,
                            商品名 = ws[f"C{s}"].value,
                            下代 = ws[f"E{s}"].value,
                            update_at = datetime.datetime.now(),
                            ))
                session.commit()
                s = s + 1
                #time.sleep(1)
            else:
                print(f"update {s :} ")
                session = Session(engine)
                product = session.query(Product).filter(Product.code == ws[f"B{s}"].value).first()
                product.code = ws[f"B{s}"].value
                product.商品名 = ws[f"C{s}"].value
                product.下代 = ws[f"E{s}"].value
                product.update_at = datetime.datetime.now()
            
                session.commit()
                s = s + 1
                #time.sleep(1)
            #save(session)

    def 過去のデータをエクセルからDBに保存する(self):
        # 転記出来るデータを持ってくる。
        
    def 売り上げ解析(self):
        # 過去データから売れそうな物を順番に並べる

# ================================================================
    
      
    
import xlrd
import xlwt
from xlutils.copy import copy
"""
#xlsxを開く
rb = xlrd.open_workbook('sample.xlsx',formatting_info=True)

#シート名表示
print(rb.sheet_names())

#シート番号か、シート名で指定
sheet = sheet_by_index(0)
sheet = rb.sheet_by_name('sheet1')

#セルの値を取得、Cellの属性、もしくは、オブジェクトのメソッドCell_Valueで取得できる
cell = sheet.cell(1,2).value
cell = sheet.cell_value(1, 2)

#列指定で取得(xlrd.sheet.Cellで取れる)
col = sheet.col(1)

#列指定の値取得(Listで取れる)
col_values = sheet.col_values(1)

#行指定で取得
row = sheet.row(1)

#行指定の値取得(Listで取れる)
row_values = sheet.row_values(1)

#二次元配列で取り出し
def get_list_2d_all(sheet):
    return [sheet.row_values(row) for row in range(sheet.nrows)]

l_2d_all = get_list_2d_all(sheet)
print(l_2d_all[1][0])

#テンプレートがある場合はxlutilsで書式を複製する
wb = copy(rb)

#新しいシート追加
sheet = wb.add_sheet('sheet3')

#sheet.write(行, 列, '値')で書き込む
sheet.write(0, 0, 'A')
sheet.write(0, 1, 'B')
sheet.write(1, 0, 10)
sheet.write(1, 1, 20)

#書き込み
wb.save('xlwt_sample.xlsx')
"""


if __name__ == '__main__':
    
    excel_analysis = Excel_Analysis()
    excel_analysis.excel_slise()