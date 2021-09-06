# -*- coding:utf-8 -*-

# 2021/04/25
# python -m pip install --upgrade pip
# pip install -U sqlalchemy
# pip install -U flake8
# pip install -U autopep8

# pip install -U openpyxl
# pip install -U beautifulsoup4
# pip install -U requests
# pip install -U pyinstaller

# pyinstaller kabu.py --onefile

from openpyxl import load_workbook, Workbook

from config import Config

import datetime
import time
import os
import random
import re
import glob

from sqlalchemy.ext.automap import automap_base
from sqlalchemy.orm import Session
from sqlalchemy import create_engine

Base = automap_base()

# engine, suppose it has two tables 'user' and 'address' set up
#engine = create_engine("sqlite:///db.sqlite3")

# reflect the tables
#Base.prepare(engine, reflect=True)

# mapped classes are now created with names by default
# matching that of the table name.
#affiliate_video = Base.classes.Affiliate_Video_affiliate_video
#affiliate_video_pic = Base.classes.Affiliate_Video_affiliate_video_pic

#session = Session(engine)


class Excel_db():
    def setup(self):
        Base = automap_base()

        self.config = Config()

        # engine, suppose it has two tables 'user' and 'address' set up
        self.engine = create_engine(self.config.db_path)
        # 個人情報の塊なのでGITの外に設置

        # reflect the tables
        Base.prepare(self.engine, reflect=True)

        # mapped classes are now created with names by default
        # matching that of the table name.
        
        self.seeds = Base.classes.seeds
        self.seed = Base.classes.seed
        self.site = Base.classes.site
        self.sale = Base.classes.sale
        
        self.siteID = 0
        # 0	アマゾン
        # 1	楽天
        # 2	ヤフオク
        # 3	ラクマ
        # 4	ストアーズ
        
        #self.seed = self.get
    
    def excel(self):
        session = Session(self.engine)
        seed = session.query(self.seed).all()
        
        filename = (self.config.excel_db_path)

        #parser = argparse.ArgumentParser(description="売上集計するプログラム")
        #print(os.path.exists(filename))
        if os.path.exists(filename) == False:
            print("ファイルが無いので作ります。")
            wb = Workbook(filename)
            wb.save(filename)
            wb = load_workbook(filename)
            ws = wb.active
            ws.title = "data"
            wb.save(filename)

        wb = load_workbook(filename)

        ws = wb["data"]
        s = 1
        for i in seed:
            ws[f"A{s}"] = i.code
            ws[f"B{s}"] = i.product_code
            ws[f"C{s}"] = i.Product_name
            ws[f"D{s}"] = i.quantity
            
            ws[f"E{s}"] = i.buyer_name
            ws[f"F{s}"] = i.buyer_pen_name
            ws[f"G{s}"] = i.buyer_zip_code
            ws[f"H{s}"] = i.buyer_address
            ws[f"I{s}"] = i.buyer_phone_number
            ws[f"J{s}"] = i.buyer_email_address
            
            ws[f"K{s}"] = i.destination_name
            ws[f"L{s}"] = i.destination_pen_name
            ws[f"M{s}"] = i.destination_zip_code
            ws[f"N{s}"] = i.destination_address
            ws[f"O{s}"] = i.destination_phone_number
            ws[f"P{s}"] = i.destination_email_address
            
            ws[f"Q{s}"] = i.purchase_price
            ws[f"R{s}"] = i.shipping
            ws[f"S{s}"] = i.total_fee
            
            s = s + 1

        wb.save(self.config.excel_db_path)

if __name__ == '__main__':
    excel_db = Excel_db()
    excel_db.setup()
    excel_db.excel()
    #print(mail_order_site.siteID)
    #print(mail_order_site.get_seed())
