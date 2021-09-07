import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import colors ,Border, Side, PatternFill, Font, Color ,GradientFill, Alignment, numbers
from openpyxl.chart import (
    LineChart,
    Reference,
)

#既存ファイルを開く
WorkBook = openpyxl.load_workbook('sample2.xlsx')

#新規作成
#WorkBook = openpyxl.Workbook()

#シート名を取得(List型)
print(WorkBook.sheetnames)
CheckName = WorkBook.sheetnames[0]

#アクティブなbookから、アクティブなシートを指定
sheet = WorkBook.active

# 名前を指定
sheet = WorkBook.worksheets['Sheet_New']

# 新規作成(0番目に配置)      
sheet = WorkBook.create_sheet('Sheet_New', 0)

# 既存からコピー 
sheet = WorkBook.copy_worksheet('Sheet_1')

#シート名で指定
if CheckName == 'Sheet_1':
    sheet = WorkBook['Sheet_1']
if CheckName == 'Sheet_XX':
    sheet = WorkBook['Sheet_XX']

#シート名を書き換え
sheet.title = "Sheet_XX"

#セルを指定して読み出し
print(sheet["B4"].value)
print(sheet.cell(column=10, row=10).value)

#セルに値を入れる
sheet["B4"] = "ABCDE"

#行と列を指定しても書き込める
sheet.cell(row = 1, column = 1, value = 5)
sheet.cell(row = 2, column = 1, value = 10)
sheet.cell(row = 3, column = 1, value = 15)
sheet.cell(row = 2, column = 3, value = 23)
sheet.cell(row = 4, column = 6, value = "This is F4")

#書式設定
sheet["C3"].value = 1000.99
sheet["C3"].number_format = '#,##0'
sheet["C4"].value = 100
sheet["C4"].number_format = '0.000'

# 行の高さを変更
sheet.row_dimensions[1].height = 30
# 列の幅を変更
sheet.column_dimensions['B'].width = 50

#関数を入力
sheet['B1'].value = '=SUM(A1:A3)'
sheet['B2'].value = '=AVERAGE(A:A)'

#画像貼り付け
img = Image('figure.png')
sheet.add_image(img,'D1')

#グラフ挿入
#データ範囲(対象シート,最小値列,最小値行,最大値列,最大値行)
data = Reference(sheet, min_col=2, min_row=1, max_col=4, max_row=7)
LineFig1 = LineChart()
LineFig1.title = "折れ線グラフ"
LineFig1.style = 13 # グラフ種別
LineFig1.y_axis.title = '縦タイトル'
LineFig1.x_axis.title = '横タイトル'
LineFig1Style1 = LineFig1.series[0] # 折れ線1の色
LineFig1Style1.marker.symbol = "triangle" # マーカーの形
LineFig1Style1.marker.graphicalProperties.solidFill = "FF0000" # 影色 
LineFig1Style1.marker.graphicalProperties.line.solidFill = "FF0000" # 線色
LineFig1Style1.graphicalProperties.line.noFill = True
LineFig1Style2 = LineFig1.series[1] # 折れ線2の色
LineFig1Style2.graphicalProperties.line.solidFill = "00AAAA"
LineFig1Style2.graphicalProperties.line.dashStyle = "sysDot"
LineFig1Style2.graphicalProperties.line.width = 100050
#描画
LineFig1.add_data(data, titles_from_data=True)
sheet.add_chart(LineFig1, "F1")

#指定列前に空列挿入(開始列,挿入列数)
sheet.insert_cols(7,2)

#指定行前に空行挿入(開始列,挿入行数)
sheet.insert_rows(7,2)

#指定列を削除(開始列,削除列数)
sheet.delete_cols(6,3)

#指定行を削除(開始行,削除行数)
sheet.delete_rows(6,3)

#フォント設定(フォント/書体/色/下線/塗りつぶし)
sheet['A1'].font = Font(color=colors.RED, bold=True)
sheet['A2'].font = Font(color=colors.BLUE, italic=True)
sheet['A3'].font = Font(color=colors.GREEN, name='Arial', size=14, underline="single")

#セルの水平中寄せ、垂直中央寄せ
sheet['C2'].value = 'TEXT'
sheet['C2'].alignment = Alignment(horizontal="center", vertical="center")

#セル結合
sheet.merge_cells('A2:D2')
#セル結合解除
sheet.unmerge_cells('A2:D2')

#背景色指定
sheet['A3'].fill = PatternFill(patternType='solid', fgColor='FDE9D9')

#罫線(Cellに対して上下左右)
#線種指定(‘dashDot’, ‘dashDotDot’, ‘double’, ‘hair’, ‘dotted’, ‘mediumDashDotDot’, ‘dashed’, ‘mediumDashed’, ‘slantDashDot’, ‘thick’, ‘thin’, ‘medium’, ‘mediumDashDot’)
BlackLine = Side(style='thin', color='000000')
BlueLineThick = Side(style='thick', color='0000FF')

#セルに対して指定
sheet['B1'].border = Border(top= BlackLine, bottom= BlueLineThick, left= BlackLine, right= BlackLine)

#印刷用紙設定
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_TABLOID

#余白設定(inch)
sheet.page_setup.fitToHeight = 0
sheet.page_setup.fitToWidth = 1

#印刷位置 : 中寄せ
sheet.print_options.horizontalCentered = True
sheet.print_options.verticalCentered = True

#印刷ヘッダー/フッター[全ページ共通] (左寄せ:left,右寄せ:right,中寄せ:center)
sheet.oddHeader.left.text = "Page &[Page] of &N"
sheet.oddHeader.left.size = 14
sheet.oddHeader.left.font = "MS Gothic,Bold"
sheet.oddHeader.left.color = "CC3366"
sheet.oddFooter.center.text = "Page &[Page] of &N"

#印刷ヘッダー/フッダー[1ページ目のみ]
sheet.firstHeader.center.text = "Title"
sheet.firstFooter.center.text = "Page &[Page] of &N"

#印刷ヘッダー/フッダー[1ページ目以外]
sheet.evenHeader.center.text = "Title"
sheet.evenFooter.center.text = "Page &[Page] of &N"

#印刷タイトル行、列指定
sheet.print_title_cols = 'A:B' 
sheet.print_title_rows = '1:1' 

#印刷範囲
sheet.print_area = 'A1:F10'

# 保存
WorkBook.save('sample2.xlsx')